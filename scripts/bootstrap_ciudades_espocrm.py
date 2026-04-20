#!/usr/bin/env python3
"""
bootstrap_ciudades_espocrm.py — Ejecución única (idempotente)

Crea la entidad Ciudad en EspoCRM y la puebla con todas las ciudades
de la plantilla de importación de clientes Effi.

Qué hace:
  1. Crea entityDefs/Ciudad.json + scopes/Ciudad.json + i18n/Ciudad.json
  2. Agrega campo 'ciudad' (link) en Contact entityDefs, i18n y layout
  3. Corre rebuild.php para crear la tabla `ciudad` en DB y agregar `ciudad_id` en contact
  4. Inserta todas las ciudades de la plantilla en la tabla `ciudad` de espocrm
     (idempotente: ignora duplicados por id_effi)

Fuente de datos:
  plantilla_importacion_cliente.xlsx (raíz del proyecto)
  Hojas: Ciudades Colombia, Ciudades Ecuador, Ciudades Rep Dominicana, Ciudades Guatemala

Ejecutar manualmente:
  python3 scripts/bootstrap_ciudades_espocrm.py
"""

import json
import secrets
import subprocess
import sys
from pathlib import Path

import mysql.connector
import openpyxl

# ─── Configuración ─────────────────────────────────────────────────────────────

PLANTILLA = Path(__file__).parent.parent / 'plantilla_importacion_cliente.xlsx'

import os as _os, sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
from lib import cfg_local
DB_ESPO = dict(**cfg_local(), database='espocrm')

ESPOCRM_CONTAINER = 'espocrm'

# Rutas dentro del contenedor
ENTITY_DEFS_CIUDAD  = '/var/www/html/custom/Espo/Custom/Resources/metadata/entityDefs/Ciudad.json'
SCOPES_CIUDAD       = '/var/www/html/custom/Espo/Custom/Resources/metadata/scopes/Ciudad.json'
I18N_CIUDAD         = '/var/www/html/custom/Espo/Custom/Resources/i18n/es_MX/Ciudad.json'
ENTITY_DEFS_CONTACT = '/var/www/html/custom/Espo/Custom/Resources/metadata/entityDefs/Contact.json'
I18N_CONTACT        = '/var/www/html/custom/Espo/Custom/Resources/i18n/es_MX/Contact.json'
LAYOUT_CONTACT      = '/var/www/html/custom/Espo/Custom/Resources/layouts/Contact/detail.json'
REBUILD_SCRIPT      = '/var/www/html/rebuild.php'
CLEAR_CACHE_SCRIPT  = '/var/www/html/clear_cache.php'

# Archivos temporales
TMP_ENTITY_CIUDAD  = '/tmp/espo_ciudad_entitydefs.json'
TMP_SCOPES_CIUDAD  = '/tmp/espo_ciudad_scopes.json'
TMP_I18N_CIUDAD    = '/tmp/espo_ciudad_i18n.json'
TMP_ENTITY_CONTACT = '/tmp/espo_contact_entitydefs_v2.json'
TMP_I18N_CONTACT   = '/tmp/espo_contact_i18n_v2.json'
TMP_LAYOUT_CONTACT = '/tmp/espo_contact_layout_v2.json'

# Hojas con ciudades en la plantilla (nombre hoja → país)
HOJAS_CIUDADES = {
    'Ciudades Colombia':       'Colombia',
    'Ciudades Ecuador':        'Ecuador',
    'Ciudades Rep Dominicana': 'República Dominicana',
    'Ciudades Guatemala':      'Guatemala',
}

# ─── Helpers ───────────────────────────────────────────────────────────────────

def run(cmd, check=True):
    result = subprocess.run(cmd, capture_output=True, text=True)
    if check and result.returncode != 0:
        raise RuntimeError(f"Comando falló: {' '.join(cmd)}\n{result.stderr}")
    return result.stdout.strip()


def gen_id():
    return secrets.token_hex(9)[:17]


def docker_read_json(path):
    """Lee un JSON del contenedor EspoCRM."""
    raw = run(['docker', 'exec', ESPOCRM_CONTAINER, 'cat', path], check=False)
    if not raw:
        return {}
    return json.loads(raw)


# ─── Paso 1: Crear entityDefs, scopes e i18n de Ciudad ─────────────────────────

def crear_metadata_ciudad():
    entity_defs = {
        'fields': {
            'name': {
                'type': 'varchar',
                'required': True,
                'maxLength': 200
            },
            'idEffi': {
                'type': 'int'
            },
            'departamento': {
                'type': 'varchar',
                'maxLength': 150
            },
            'pais': {
                'type': 'varchar',
                'maxLength': 100
            }
        },
        'links': {},
        'collection': {
            'orderBy': 'name',
            'order': 'asc',
            'textFilterFields': ['name', 'departamento', 'pais']
        },
        'indexes': {
            'idEffi': {
                'unique': True,
                'columns': ['id_effi']
            }
        }
    }

    scopes = {
        'entity': True,
        'importable': False,
        'exportFormatList': [],
        'duplicateCheckFieldList': ['idEffi']
    }

    i18n = {
        'fields': {
            'name':         'Nombre Ciudad',
            'idEffi':       'ID Effi',
            'departamento': 'Departamento',
            'pais':         'País'
        },
        'labels': {
            'Ciudad': 'Ciudad'
        }
    }

    with open(TMP_ENTITY_CIUDAD, 'w', encoding='utf-8') as f:
        json.dump(entity_defs, f, ensure_ascii=False, indent=2)
    with open(TMP_SCOPES_CIUDAD, 'w', encoding='utf-8') as f:
        json.dump(scopes, f, ensure_ascii=False, indent=2)
    with open(TMP_I18N_CIUDAD, 'w', encoding='utf-8') as f:
        json.dump(i18n, f, ensure_ascii=False, indent=2)


# ─── Paso 2: Actualizar Contact para agregar campo ciudad (link) ────────────────

def actualizar_contact_metadata():
    # Leer Contact.json actual (ya tiene los campos custom anteriores)
    contact_defs = docker_read_json(ENTITY_DEFS_CONTACT)

    # Agregar campo ciudad (link) si no existe
    if 'ciudad' not in contact_defs.get('fields', {}):
        contact_defs.setdefault('fields', {})['ciudad'] = {
            'type': 'link',
            'entity': 'Ciudad'
        }

    # Agregar link en la sección links
    if 'ciudad' not in contact_defs.get('links', {}):
        contact_defs.setdefault('links', {})['ciudad'] = {
            'type': 'belongsTo',
            'entity': 'Ciudad'
        }

    # Leer i18n Contact actual
    contact_i18n = docker_read_json(I18N_CONTACT)
    contact_i18n.setdefault('fields', {})['ciudad'] = 'Ciudad'
    contact_i18n.setdefault('links', {})['ciudad'] = 'Ciudad'

    # Layout: leer actual y agregar ciudad si no está
    layout = docker_read_json(LAYOUT_CONTACT)
    if not layout:
        layout = [{'label': '', 'rows': []}]

    # Verificar que ciudad no esté ya en el layout
    ya_tiene_ciudad = any(
        any(cell and isinstance(cell, dict) and cell.get('name') == 'ciudad'
            for cell in row)
        for panel in layout
        for row in panel.get('rows', [])
    )

    if not ya_tiene_ciudad:
        # Insertar ciudad después del campo address (penúltima fila antes de description)
        for panel in layout:
            rows = panel.get('rows', [])
            # Buscar la posición de address
            for i, row in enumerate(rows):
                if any(isinstance(cell, dict) and cell.get('name') == 'address' for cell in row):
                    # Insertar ciudad antes de address
                    rows.insert(i, [{'name': 'ciudad'}, False])
                    break

    with open(TMP_ENTITY_CONTACT, 'w', encoding='utf-8') as f:
        json.dump(contact_defs, f, ensure_ascii=False, indent=2)
    with open(TMP_I18N_CONTACT, 'w', encoding='utf-8') as f:
        json.dump(contact_i18n, f, ensure_ascii=False, indent=2)
    with open(TMP_LAYOUT_CONTACT, 'w', encoding='utf-8') as f:
        json.dump(layout, f, ensure_ascii=False, indent=2)


# ─── Paso 3: Aplicar metadata y rebuild ────────────────────────────────────────

def aplicar_metadata():
    # Crear directorios necesarios en el contenedor
    for d in [
        '/var/www/html/custom/Espo/Custom/Resources/metadata/entityDefs',
        '/var/www/html/custom/Espo/Custom/Resources/metadata/scopes',
        '/var/www/html/custom/Espo/Custom/Resources/i18n/es_MX',
        '/var/www/html/custom/Espo/Custom/Resources/layouts/Contact',
    ]:
        run(['docker', 'exec', ESPOCRM_CONTAINER, 'mkdir', '-p', d])

    # Ciudad
    run(['docker', 'cp', TMP_ENTITY_CIUDAD, f'{ESPOCRM_CONTAINER}:{ENTITY_DEFS_CIUDAD}'])
    run(['docker', 'cp', TMP_SCOPES_CIUDAD, f'{ESPOCRM_CONTAINER}:{SCOPES_CIUDAD}'])
    run(['docker', 'cp', TMP_I18N_CIUDAD,   f'{ESPOCRM_CONTAINER}:{I18N_CIUDAD}'])

    # Contact actualizado
    run(['docker', 'cp', TMP_ENTITY_CONTACT, f'{ESPOCRM_CONTAINER}:{ENTITY_DEFS_CONTACT}'])
    run(['docker', 'cp', TMP_I18N_CONTACT,   f'{ESPOCRM_CONTAINER}:{I18N_CONTACT}'])
    run(['docker', 'cp', TMP_LAYOUT_CONTACT, f'{ESPOCRM_CONTAINER}:{LAYOUT_CONTACT}'])

    # Rebuild
    run(['docker', 'exec', ESPOCRM_CONTAINER, 'php', REBUILD_SCRIPT])
    run(['docker', 'exec', ESPOCRM_CONTAINER, 'php', CLEAR_CACHE_SCRIPT])


# ─── Paso 4: Poblar tabla ciudad ───────────────────────────────────────────────

def poblar_ciudades():
    if not PLANTILLA.exists():
        raise FileNotFoundError(f'No se encontró la plantilla: {PLANTILLA}')

    wb = openpyxl.load_workbook(str(PLANTILLA), data_only=True, read_only=True)

    conn = mysql.connector.connect(**DB_ESPO, autocommit=False)
    cur  = conn.cursor()

    now = '2026-01-01 00:00:00'
    total = 0
    skip  = 0

    for hoja, pais in HOJAS_CIUDADES.items():
        if hoja not in wb.sheetnames:
            print(f'   ⚠ Hoja "{hoja}" no encontrada, omitiendo')
            continue

        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        # Columnas: (Nombre departamento, ID EFFI ciudad, Nombre ciudad)
        for fila in filas[1:]:  # saltar encabezado
            if not fila or fila[1] is None or fila[2] is None:
                skip += 1
                continue

            departamento = str(fila[0]).strip() if fila[0] else ''
            id_effi      = int(fila[1])
            nombre       = str(fila[2]).strip()

            if not nombre or not id_effi:
                skip += 1
                continue

            espo_id = gen_id()
            try:
                cur.execute(
                    """INSERT IGNORE INTO ciudad
                       (id, deleted, name, id_effi, departamento, pais)
                       VALUES (%s, 0, %s, %s, %s, %s)""",
                    (espo_id, nombre, id_effi, departamento, pais)
                )
                total += cur.rowcount
            except Exception as e:
                print(f'   ⚠ Error insertando {nombre}: {e}')
                skip += 1

        conn.commit()

    cur.close()
    conn.close()
    wb.close()
    return total, skip


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    print('▶ bootstrap_ciudades_espocrm.py')

    print('  [1/4] Generando metadata de entidad Ciudad...')
    crear_metadata_ciudad()

    print('  [2/4] Actualizando Contact (campo ciudad link)...')
    actualizar_contact_metadata()

    print('  [3/4] Aplicando metadata + rebuild EspoCRM...')
    aplicar_metadata()
    print('       ✅ Rebuild completado — tabla ciudad creada')

    print('  [4/4] Poblando ciudades desde la plantilla...')
    total, skip = poblar_ciudades()
    print(f'       ✅ {total} ciudades insertadas, {skip} omitidas')

    print('✅ bootstrap_ciudades_espocrm — completo')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'❌ bootstrap_ciudades_espocrm — ERROR: {e}', file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
